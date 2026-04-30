"""
Polymarket Wallet Tracker & Market Predictor
Tracks wallet trade history and predicts next market investments.

Requirements:
    pip install requests pandas openpyxl python-dotenv schedule

Usage:
    python polymarket_tracker.py

Railway setup:
    - Mount a persistent volume at /data
    - Set OUTPUT_DIR=/data in environment variables
"""

import os
import json
import time
import requests
import pandas as pd
from datetime import datetime, timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter
import schedule

# ── Config ────────────────────────────────────────────────────────────────────

WALLETS = [
    "0x53e55bc7cb3d67ad177c023ce891ad076a9d6177",
    "0xc6587b11a2209e46dfe3928b31c5514a8e33b784",
]

OUTPUT_DIR = os.environ.get("OUTPUT_DIR", "/data")
EXCEL_PATH = os.path.join(OUTPUT_DIR, "polymarket_tracker.xlsx")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Polymarket endpoints
CLOB_API       = "https://clob.polymarket.com"
GAMMA_API      = "https://gamma-api.polymarket.com"
SUBGRAPH_URL   = "https://api.thegraph.com/subgraphs/name/polymarket/matic-markets"

HEADERS = {"Accept": "application/json"}

# ── Data Fetching ──────────────────────────────────────────────────────────────

def fetch_trades_clob(wallet: str) -> list[dict]:
    """Fetch trade history from Polymarket CLOB API."""
    trades = []
    next_cursor = None

    while True:
        params = {"maker_address": wallet.lower(), "limit": 500}
        if next_cursor:
            params["next_cursor"] = next_cursor

        try:
            r = requests.get(f"{CLOB_API}/trades", params=params, headers=HEADERS, timeout=15)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"  [CLOB] Error fetching trades for {wallet}: {e}")
            break

        batch = data.get("data", [])
        trades.extend(batch)

        next_cursor = data.get("next_cursor")
        if not next_cursor or next_cursor == "LTE=":
            break
        time.sleep(0.3)

    return trades


def fetch_positions_gamma(wallet: str) -> list[dict]:
    """Fetch open/closed positions from Gamma API."""
    try:
        r = requests.get(
            f"{GAMMA_API}/positions",
            params={"user": wallet.lower(), "limit": 500},
            headers=HEADERS,
            timeout=15,
        )
        r.raise_for_status()
        return r.json() if isinstance(r.json(), list) else r.json().get("data", [])
    except Exception as e:
        print(f"  [Gamma] Error fetching positions for {wallet}: {e}")
        return []


def fetch_market_info(condition_id: str) -> dict:
    """Fetch market metadata from Gamma API."""
    try:
        r = requests.get(
            f"{GAMMA_API}/markets",
            params={"condition_id": condition_id},
            headers=HEADERS,
            timeout=10,
        )
        r.raise_for_status()
        data = r.json()
        if isinstance(data, list) and data:
            return data[0]
        elif isinstance(data, dict):
            return data.get("data", [{}])[0] if data.get("data") else {}
    except Exception:
        pass
    return {}


# ── Analysis ───────────────────────────────────────────────────────────────────

def parse_trades(raw_trades: list[dict], wallet: str) -> pd.DataFrame:
    """Normalize raw trade data into a clean DataFrame."""
    rows = []
    for t in raw_trades:
        try:
            rows.append({
                "wallet":        wallet,
                "trade_id":      t.get("id", ""),
                "timestamp":     datetime.fromtimestamp(
                                     int(t.get("created_at", 0) or 0), tz=timezone.utc
                                 ),
                "market_id":     t.get("market", ""),
                "condition_id":  t.get("condition_id", ""),
                "side":          t.get("side", "").upper(),           # BUY / SELL
                "outcome":       t.get("outcome", ""),               # YES / NO
                "price":         float(t.get("price", 0) or 0),      # 0-1 scale
                "size":          float(t.get("size", 0) or 0),       # USDC
                "status":        t.get("status", ""),
                "asset_id":      t.get("asset_id", ""),
            })
        except Exception:
            continue
    df = pd.DataFrame(rows)
    if not df.empty:
        df.sort_values("timestamp", ascending=False, inplace=True)
    return df


def enrich_with_market_names(df: pd.DataFrame) -> pd.DataFrame:
    """Add market question/name column by looking up unique condition IDs."""
    if df.empty:
        return df

    unique_cids = df["condition_id"].dropna().unique()
    cid_to_name = {}

    print(f"  Fetching metadata for {len(unique_cids)} markets...")
    for cid in unique_cids:
        if not cid:
            continue
        info = fetch_market_info(cid)
        cid_to_name[cid] = info.get("question") or info.get("title") or cid[:20]
        time.sleep(0.2)

    df["market_name"] = df["condition_id"].map(cid_to_name).fillna("Unknown")
    return df


def compute_wallet_stats(df: pd.DataFrame) -> dict:
    """Compute summary statistics per wallet."""
    if df.empty:
        return {}

    buys  = df[df["side"] == "BUY"]
    sells = df[df["side"] == "SELL"]

    return {
        "total_trades":    len(df),
        "total_buy_vol":   buys["size"].sum(),
        "total_sell_vol":  sells["size"].sum(),
        "unique_markets":  df["condition_id"].nunique(),
        "avg_buy_price":   buys["price"].mean() if not buys.empty else 0,
        "avg_sell_price":  sells["price"].mean() if not sells.empty else 0,
        "first_trade":     df["timestamp"].min(),
        "last_trade":      df["timestamp"].max(),
        "fav_outcome":     df["outcome"].mode()[0] if not df["outcome"].empty else "N/A",
    }


def predict_next_markets(df: pd.DataFrame, top_n: int = 5) -> pd.DataFrame:
    """
    Heuristic prediction of next likely markets and entry prices.

    Signals used:
      1. Category/tag frequency — wallets tend to specialise
      2. Recency-weighted market revisit probability
      3. Average entry price per market → likely next entry range
      4. Win-rate on outcome side (YES/NO preference)
    """
    if df.empty:
        return pd.DataFrame()

    # Recency weight: exponential decay over days
    now = datetime.now(tz=timezone.utc)
    df = df.copy()
    df["days_ago"] = (now - df["timestamp"]).dt.total_seconds() / 86400
    df["weight"]   = df["days_ago"].apply(lambda d: 0.95 ** d)

    # Weighted score per market
    mkt_scores = (
        df.groupby(["condition_id", "market_name"])
        .agg(
            total_weight    = ("weight", "sum"),
            trade_count     = ("trade_id", "count"),
            avg_buy_price   = ("price", lambda x: x[df.loc[x.index, "side"] == "BUY"].mean() if (df.loc[x.index, "side"] == "BUY").any() else None),
            last_trade_date = ("timestamp", "max"),
            net_usdc        = ("size", lambda x: x[df.loc[x.index, "side"] == "BUY"].sum() - x[df.loc[x.index, "side"] == "SELL"].sum()),
            yes_pct         = ("outcome", lambda x: (x == "YES").mean()),
        )
        .reset_index()
    )

    # Markets still open (net positive USDC in = more bought than sold)
    open_positions  = mkt_scores[mkt_scores["net_usdc"] > 5].copy()
    closed_markets  = mkt_scores[mkt_scores["net_usdc"] <= 5].copy()

    # Predict: frequently re-entered closed markets are likely candidates
    closed_markets["prediction_score"] = (
        closed_markets["total_weight"] * 0.6 +
        closed_markets["trade_count"]  * 0.4
    )
    predictions = closed_markets.sort_values("prediction_score", ascending=False).head(top_n)

    predictions["predicted_entry_odds"] = predictions["avg_buy_price"].apply(
        lambda p: f"{p:.2f} ({p*100:.0f}¢)" if pd.notna(p) else "N/A"
    )
    predictions["likely_outcome"] = (predictions["yes_pct"] >= 0.5).map({True: "YES", False: "NO"})

    return predictions[[
        "market_name", "condition_id", "trade_count",
        "last_trade_date", "predicted_entry_odds",
        "likely_outcome", "prediction_score"
    ]].rename(columns={
        "market_name":          "Market",
        "condition_id":         "Condition ID",
        "trade_count":          "Historical Trades",
        "last_trade_date":      "Last Seen",
        "predicted_entry_odds": "Predicted Entry (odds)",
        "likely_outcome":       "Likely Outcome",
        "prediction_score":     "Confidence Score",
    })


# ── Excel Output ───────────────────────────────────────────────────────────────

BLUE   = "FF003087"
GOLD   = "FFFFD700"
LGREEN = "FFD9F0D9"
WHITE  = "FFFFFFFF"
LGRAY  = "FFF2F2F2"

def _header_style(cell, bg=BLUE, fg=WHITE):
    cell.font      = Font(bold=True, color=fg, size=11)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _thin_border():
    s = Side(style="thin", color="FFCCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _write_df_to_sheet(ws, df: pd.DataFrame, start_row=1):
    """Write a DataFrame to a worksheet with styling."""
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=ci, value=col)
        _header_style(cell)
        ws.column_dimensions[get_column_letter(ci)].width = max(18, len(str(col)) + 4)

    for ri, row in enumerate(df.itertuples(index=False), start_row + 1):
        fill_color = LGRAY if ri % 2 == 0 else WHITE
        for ci, val in enumerate(row, 1):
            # Format datetimes
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d %H:%M UTC")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = PatternFill("solid", start_color=fill_color)
            cell.border    = _thin_border()
            cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.auto_filter.ref = ws.dimensions


def build_excel(all_trades: dict, all_predictions: dict, wallet_stats: dict):
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.sheet_view.showGridLines = False

    ws_sum["A1"] = "Polymarket Wallet Intelligence Report"
    ws_sum["A1"].font = Font(bold=True, size=16, color=BLUE[2:])
    ws_sum["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}"
    ws_sum["A2"].font = Font(italic=True, size=10)

    row = 4
    for wallet, stats in wallet_stats.items():
        ws_sum.cell(row=row, column=1, value=f"Wallet: {wallet}").font = Font(bold=True, size=12)
        row += 1
        for k, v in stats.items():
            label_cell = ws_sum.cell(row=row, column=1, value=k.replace("_", " ").title())
            label_cell.font = Font(bold=True)
            val = v.strftime("%Y-%m-%d %H:%M UTC") if isinstance(v, datetime) else v
            ws_sum.cell(row=row, column=2, value=round(val, 4) if isinstance(val, float) else val)
            row += 1
        row += 1

    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 35

    # ── Per-wallet trade history + predictions ─────────────────────────────────
    for wallet in WALLETS:
        short = wallet[:8]

        # Trade history
        df = all_trades.get(wallet, pd.DataFrame())
        ws_t = wb.create_sheet(f"Trades {short}")
        if not df.empty:
            display_cols = ["timestamp", "market_name", "side", "outcome",
                            "price", "size", "status", "condition_id"]
            display_cols = [c for c in display_cols if c in df.columns]
            _write_df_to_sheet(ws_t, df[display_cols])
        else:
            ws_t["A1"] = "No trades found"

        # Predictions
        preds = all_predictions.get(wallet, pd.DataFrame())
        ws_p = wb.create_sheet(f"Predictions {short}")
        ws_p["A1"] = f"Next Market Predictions — {wallet}"
        ws_p["A1"].font = Font(bold=True, size=13, color=BLUE[2:])
        ws_p["A2"] = (
            "Confidence Score = recency-weighted trade frequency. "
            "Higher = more likely next entry."
        )
        ws_p["A2"].font = Font(italic=True, size=9)

        if not preds.empty:
            _write_df_to_sheet(ws_p, preds, start_row=4)
            # Highlight top prediction
            for ci in range(1, len(preds.columns) + 1):
                cell = ws_p.cell(row=5, column=ci)
                cell.fill = PatternFill("solid", start_color=GOLD)
                cell.font = Font(bold=True)
        else:
            ws_p["A4"] = "Insufficient data for predictions"

    wb.save(EXCEL_PATH)
    print(f"\n✅ Excel saved → {EXCEL_PATH}")


# ── Main ───────────────────────────────────────────────────────────────────────

def run():
    print(f"\n{'='*60}")
    print(f"  Polymarket Tracker — {datetime.now().strftime('%Y-%m-%d %H:%M')} UTC")
    print(f"{'='*60}")

    all_trades      = {}
    all_predictions = {}
    wallet_stats    = {}

    for wallet in WALLETS:
        print(f"\n📍 Wallet: {wallet}")

        print("  Fetching CLOB trade history...")
        raw = fetch_trades_clob(wallet)
        print(f"  → {len(raw)} raw trades")

        df = parse_trades(raw, wallet)
        df = enrich_with_market_names(df)

        all_trades[wallet]   = df
        wallet_stats[wallet] = compute_wallet_stats(df)

        print("  Running prediction model...")
        preds = predict_next_markets(df, top_n=10)
        all_predictions[wallet] = preds

        if not preds.empty:
            print(f"\n  🔮 Top 3 predicted next markets:")
            for _, r in preds.head(3).iterrows():
                print(f"     • {r['Market'][:60]}")
                print(f"       Entry odds: {r['Predicted Entry (odds)']}  |  Likely: {r['Likely Outcome']}")

    build_excel(all_trades, all_predictions, wallet_stats)


def schedule_run():
    """Run once now, then every 6 hours."""
    run()
    schedule.every(6).hours.do(run)
    print("\n⏰ Scheduled to refresh every 6 hours. Ctrl+C to stop.")
    while True:
        schedule.run_pending()
        time.sleep(60)


if __name__ == "__main__":
    mode = os.environ.get("RUN_MODE", "once")
    if mode == "scheduled":
        schedule_run()
    else:
        run()
