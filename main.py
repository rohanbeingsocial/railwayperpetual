import os
import requests
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print('✅ Ready')

# --- 3. Config ---
POLL_INTERVAL       = 60    
LEADERBOARD_REFRESH = 30    
TOP_N               = 20    
MIN_POSITION_USD    = 10    
PAPER_BALANCE       = 1000.0
BET_SIZE_PCT        = 0.05  

DATA_API   = 'https://data-api.polymarket.com/v1'
CLOB_API   = 'https://clob.polymarket.com'

# IMPORTANT: On Railway, we will map this to a persistent volume (e.g., /data/copy_trade_pnl.xlsx)
# If not on Railway, it saves in the local directory.
EXCEL_FILE = os.getenv('EXCEL_PATH', 'copy_trade_pnl.xlsx')

CATEGORIES = ['POLITICS', 'ECONOMICS', 'FINANCE', 'CULTURE', 'TECH', 'SPORTS', 'CRYPTO']

session = requests.Session()
session.headers.update({
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'application/json',
})

print('✅ Config loaded')
print(f'   Paper balance : ${PAPER_BALANCE:,.0f}')
print(f'   Bet size      : {BET_SIZE_PCT*100:.0f}% per trade')
print(f'   Watch         : top {TOP_N} per category, refreshed every {LEADERBOARD_REFRESH} cycles')
print(f'   Save Path     : {EXCEL_FILE}')

# --- 4. Functions ---
def fetch_top_wallets(top_n=TOP_N):
    wallet_map = {}
    for cat in CATEGORIES:
        try:
            res = session.get(f'{DATA_API}/leaderboard', params={'timePeriod': 'MONTH', 'limit': top_n, 'orderBy': 'PNL', 'category': cat}, timeout=15)
            if res.status_code != 200: continue
            traders = res.json()
            if not isinstance(traders, list): continue

            for t in traders:
                w = t.get('proxyWallet', '')
                if not w or not w.startswith('0x'): continue
                pnl = float(t.get('pnl', 0))

                if w not in wallet_map:
                    wallet_map[w] = {'username': t.get('userName') or w[:10], 'wallet': w, 'categories': [cat], 'pnl': pnl}
                else:
                    if cat not in wallet_map[w]['categories']: wallet_map[w]['categories'].append(cat)
                    wallet_map[w]['pnl'] = max(wallet_map[w]['pnl'], pnl)
            time.sleep(0.3)
        except Exception as e:
            print(f'  ⚠️ Leaderboard error ({cat}): {e}')

    for w in wallet_map:
        wallet_map[w]['category'] = '+'.join(wallet_map[w]['categories'])
    return wallet_map 

def build_watch_list(live_wallets, open_positions, seen_tx):
    combined = dict(live_wallets)
    locked_wallets = {pos['wallet'] for pos in open_positions.values()}
    for wallet in locked_wallets:
        if wallet not in combined:
            pos = next(p for p in open_positions.values() if p['wallet'] == wallet)
            combined[wallet] = {'username': pos['username'], 'wallet': wallet, 'category': pos['category'], 'pnl': 0, 'locked': True}
    for wallet in combined:
        seen_tx.setdefault(wallet, set())
    return combined

def get_activity(wallet, limit=20):
    try:
        res = session.get(f'{DATA_API}/activity', params={'user': wallet, 'limit': limit}, timeout=15)
        return res.json() if res.status_code == 200 else []
    except:
        return []

def get_current_price(token_id):
    try:
        res = session.get(f'{CLOB_API}/midpoint', params={'token_id': token_id}, timeout=10)
        if res.status_code == 200: return float(res.json().get('mid', 0))
    except: pass
    return None

def parse_trade(item):
    try:
        price = float(item.get('price', 0))
        size  = float(item.get('size', 0))
        return {
            'market': item.get('title') or item.get('market', 'Unknown'),
            'outcome': item.get('outcome', '?'), 'side': item.get('side', '?'),
            'price': round(price, 4), 'prob_pct': round(price * 100, 1),
            'size': round(size, 2), 'usd_value': round(price * size, 2),
            'tx_hash': item.get('transactionHash', ''), 'timestamp': item.get('timestamp', 0),
            'asset': item.get('asset', ''), 'condition_id': item.get('conditionId', ''),
            'outcome_index': item.get('outcomeIndex', -1),
        }
    except: return None

def pos_key(wallet, condition_id, outcome_index):
    return f'{wallet[:10]}_{condition_id[:12]}_{outcome_index}'

# --- EXCEL FORMATTING ---
OPEN_COLS = ['Opened At', 'Trader', 'Category', 'Market', 'Outcome', 'Wallet Entry', 'Our Entry', 'Entry Slip', 'Wallet Shares', 'Our Shares', 'Our USD In', 'Status', 'Asset Token', 'Wallet']
CLOSED_COLS = ['Opened At', 'Closed At', 'Trader', 'Category', 'Market', 'Outcome', 'Wallet Entry', 'Our Entry', 'Entry Slip', 'Wallet Exit', 'Our Exit', 'Exit Slip', 'Wallet P&L $', 'Our P&L $', 'P&L Diff $', 'P&L Diff %', 'Wallet ROI %', 'Our ROI %', 'Our Shares', 'Our USD In', 'Result']

def _hdr(ws, cols, bg='0F3460'):
    thin = Side(style='thin', color='CCCCCC')
    b = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = b
    ws.row_dimensions[1].height = 36

def _row(ws, r, vals, bg='FFFFFF', green_cols=None, red_cols=None):
    thin = Side(style='thin', color='CCCCCC')
    b = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_cols = green_cols or []
    red_cols   = red_cols   or []
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        if i in green_cols and isinstance(v, (int,float)) and v >= 0: c.font = Font(name='Arial', size=9, bold=True, color='1a6b2e')
        elif i in red_cols and isinstance(v, (int,float)) and v < 0: c.font = Font(name='Arial', size=9, bold=True, color='8b0000')
        elif i in green_cols or i in red_cols: c.font = Font(name='Arial', size=9, bold=True, color='1a6b2e' if (v if isinstance(v,(int,float)) else 0) >= 0 else '8b0000')
        else: c.font = Font(name='Arial', size=9)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = b
    ws.row_dimensions[r].height = 16

def init_excel():
    if os.path.exists(EXCEL_FILE):
        return # Do not overwrite if it already exists
    wb = Workbook()
    ws1 = wb.active
    ws1.title = '📂 Open Trades'
    _hdr(ws1, OPEN_COLS)
    for i, w in enumerate([20,16,20,55,10,12,12,10,12,10,11,14,42,42], 1): ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = 'A2'

    ws2 = wb.create_sheet('✅ Closed Trades')
    _hdr(ws2, CLOSED_COLS)
    for i, w in enumerate([20,20,14,18,55,10,10,10,10,10,10,10,13,13,13,12,11,11,10,11,10], 1): ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'

    ws3 = wb.create_sheet('📊 Summary')
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 22

    ws4 = wb.create_sheet('👥 Wallet Log')
    _hdr(ws4, ['Refreshed At', 'Wallet', 'Username', 'Category', 'PnL Month $', 'Status'], bg='16213E')
    for i, w in enumerate([22, 44, 20, 22, 16, 14], 1): ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.freeze_panes = 'A2'

    wb.save(EXCEL_FILE)
    print(f'📊 Excel created: {EXCEL_FILE}')

def update_excel(open_positions, closed_trades, paper_balance, watch_list):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws1 = wb['📂 Open Trades']
        for row in ws1.iter_rows(min_row=2): [setattr(c, 'value', None) for c in row]
        for i, (key, pos) in enumerate(open_positions.items(), 2):
            slip = round(pos['our_entry'] - pos['wallet_entry'], 4)
            locked = pos.get('locked', False)
            status = '🔒 LOCKED (off LB)' if locked else '✅ Active'
            _row(ws1, i, [pos['opened_at'], pos['username'], pos['category'], pos['market'], pos['outcome'], pos['wallet_entry'], pos['our_entry'], slip, pos['wallet_shares'], pos['our_shares'], pos['our_usd_in'], status, pos.get('asset',''), pos['wallet']], bg='FFF9C4' if not locked else 'FFE0B2')

        ws2 = wb['✅ Closed Trades']
        for row in ws2.iter_rows(min_row=2): [setattr(c, 'value', None) for c in row]
        for i, t in enumerate(closed_trades, 2):
            bg = 'E8F5E9' if t['our_pnl'] >= 0 else 'FFEBEE'
            _row(ws2, i, [t['opened_at'], t['closed_at'], t['username'], t['category'], t['market'], t['outcome'], t['wallet_entry'], t['our_entry'], round(t['our_entry'] - t['wallet_entry'], 4), t['wallet_exit'], t['our_exit'], round(t['our_exit'] - t['wallet_exit'], 4), round(t['wallet_pnl'], 2), round(t['our_pnl'], 2), round(t['our_pnl'] - t['wallet_pnl'], 2), round((t['our_pnl']-t['wallet_pnl'])/abs(t['wallet_pnl'])*100,1) if t['wallet_pnl'] else 0, round(t['wallet_roi']*100, 2), round(t['our_roi']*100, 2), t['our_shares'], t['our_usd_in'], '✅ WIN' if t['our_pnl'] > 0 else '❌ LOSS'], bg=bg, green_cols=[13,14,15], red_cols=[13,14,15])

        ws3 = wb['📊 Summary']
        for row in ws3.iter_rows(): [setattr(c,'value',None) for c in row]
        total_our    = sum(t['our_pnl']    for t in closed_trades)
        total_wallet = sum(t['wallet_pnl'] for t in closed_trades)
        wins         = sum(1 for t in closed_trades if t['our_pnl'] > 0)
        n = len(closed_trades)
        avg_se = sum(t['our_entry']-t['wallet_entry'] for t in closed_trades)/n if n else 0
        avg_sx = sum(t['our_exit'] -t['wallet_exit']  for t in closed_trades)/n if n else 0
        locked_count = sum(1 for p in open_positions.values() if p.get('locked'))

        rows = [('📊 SUMMARY', ''), ('Updated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')), ('', ''), ('Paper Balance', f'${paper_balance:,.2f}'), ('Wallets Watching', len(watch_list)), ('Locked (off LB)', locked_count), ('Open Positions', len(open_positions)), ('Closed Trades', n), ('', ''), ('Our Total P&L', f'${total_our:+,.2f}'), ('Wallet Total P&L', f'${total_wallet:+,.2f}'), ('Slippage Cost', f'${total_our-total_wallet:+,.2f}'), ('', ''), ('Win Rate', f'{wins/n*100:.1f}%' if n else 'N/A'), ('Wins / Losses', f'{wins} / {n-wins}'), ('', ''), ('Avg Entry Slip', f'{avg_se*100:+.2f}¢'), ('Avg Exit Slip', f'{avg_sx*100:+.2f}¢')]
        for r,(lbl,val) in enumerate(rows,1):
            ws3.cell(r,1,lbl).font = Font(name='Arial', bold=(r==1), size=10)
            ws3.cell(r,2,val).font = Font(name='Arial', size=10)
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f'⚠️ Excel error: {e}')

def log_wallet_refresh(watch_list):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws4 = wb['👥 Wallet Log']
        now = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
        r = ws4.max_row + 1
        for info in watch_list.values():
            locked = info.get('locked', False)
            _row(ws4, r, [now, info['wallet'], info['username'], info['category'], round(info['pnl'],2), '🔒 LOCKED' if locked else 'LIVE'], bg='FFF3CD' if locked else 'F0F8FF')
            r += 1
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f'⚠️ Wallet log error: {e}')

# --- ALERTS REFACTORED FOR TERMINAL ---
def alert_entry(info, trade, our_price, our_shares, our_usd):
    ts = datetime.fromtimestamp(trade['timestamp']).strftime('%H:%M:%S') if trade['timestamp'] else '?'
    slip = round((our_price - trade['price']) * 100, 2)
    print(f"\n🟢 [ENTRY COPIED] @ {ts} UTC")
    print(f"👤 {info['username']} | Market: {trade['market'][:85]}")
    print(f"   Outcome: {trade['outcome']} | Wallet Price: {trade['price']:.3f} | Our Price: {our_price:.3f} | Slip: {slip:+.2f}¢ | USD In: ${our_usd:.2f}")

def alert_exit(pos, wallet_exit, our_exit):
    wallet_pnl = (wallet_exit - pos['wallet_entry']) * pos['wallet_shares']
    our_pnl    = (our_exit    - pos['our_entry'])    * pos['our_shares']
    result = 'WIN ✅' if our_pnl >= 0 else 'LOSS ❌'
    print(f"\n{result} [POSITION CLOSED] @ {datetime.utcnow().strftime('%H:%M:%S')} UTC")
    print(f"👤 {pos['username']} | Market: {pos['market'][:65]}")
    print(f"   Wallet: Entry {pos['wallet_entry']:.3f} -> Exit {wallet_exit:.3f} | P&L: ${wallet_pnl:+.2f}")
    print(f"   Ours:   Entry {pos['our_entry']:.3f} -> Exit {our_exit:.3f} | P&L: ${our_pnl:+.2f}")

def alert_leaderboard_refresh(added, removed, locked):
    if not (added or removed): return
    print("\n🔄 LEADERBOARD REFRESHED")
    if added: print(f"   + Added: {', '.join(added[:5])}")
    if removed: print(f"   - Removed: {', '.join(removed[:5])}")
    if locked: print(f"   🔒 Kept (locked open positions): {', '.join(locked)}")

# --- 5. INITIALIZE ---
init_excel()

seen_tx        = {}
open_positions = {}
closed_trades  = []
paper_balance  = PAPER_BALANCE
watch_list     = {}

print('🌐 Fetching initial leaderboard...')
watch_list = fetch_top_wallets(TOP_N)
watch_list = build_watch_list(watch_list, open_positions, seen_tx)

print('\n📸 Snapshotting existing activity...')
for wallet, info in watch_list.items():
    activity = get_activity(wallet, limit=50)
    seen_tx[wallet] = {item.get('transactionHash','') for item in activity if item.get('transactionHash')}
    time.sleep(0.3)
log_wallet_refresh(watch_list)

# --- 6. START MONITOR ---
cycle = 0
print(f'\n🟢 LIVE | {len(watch_list)} wallets | every {POLL_INTERVAL}s | → {EXCEL_FILE}\n')

while True:
    cycle += 1
    now_str = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    events  = 0

    if cycle % LEADERBOARD_REFRESH == 0:
        prev_wallets = set(watch_list.keys())
        fresh = fetch_top_wallets(TOP_N)
        watch_list = build_watch_list(fresh, open_positions, seen_tx)
        
        new_wallets  = set(watch_list.keys())
        added   = [watch_list[w]['username'] for w in new_wallets - prev_wallets]
        removed = [w[:10] for w in prev_wallets - new_wallets]
        locked  = [watch_list[w]['username'] for w in watch_list if watch_list[w].get('locked')]

        alert_leaderboard_refresh(added, removed, locked)
        log_wallet_refresh(watch_list)

    for wallet, info in watch_list.items():
        activity = get_activity(wallet, limit=20)
        for item in activity:
            tx, side = item.get('transactionHash', ''), item.get('side', '')
            if not tx or tx in seen_tx.get(wallet, set()): continue
            seen_tx.setdefault(wallet, set()).add(tx)

            trade = parse_trade(item)
            if not trade or trade['usd_value'] < MIN_POSITION_USD: continue
            key = pos_key(wallet, trade['condition_id'], trade['outcome_index'])

            if side == 'BUY' and key not in open_positions:
                our_price  = get_current_price(trade['asset']) or trade['price']
                bet_usd    = round(paper_balance * BET_SIZE_PCT, 2)
                our_shares = round(bet_usd / our_price, 2) if our_price > 0 else 0
                paper_balance -= bet_usd

                open_positions[key] = {'username': info['username'], 'wallet': wallet, 'category': info['category'], 'market': trade['market'], 'outcome': trade['outcome'], 'opened_at': now_str, 'wallet_entry': trade['price'], 'our_entry': our_price, 'wallet_shares': trade['size'], 'our_shares': our_shares, 'our_usd_in': bet_usd, 'asset': trade['asset'], 'condition_id': trade['condition_id'], 'outcome_index': trade['outcome_index'], 'locked': info.get('locked', False)}
                events += 1
                alert_entry(info, trade, our_price, our_shares, bet_usd)
                update_excel(open_positions, closed_trades, paper_balance, watch_list)

            elif side == 'SELL' and key in open_positions:
                pos = open_positions.pop(key)
                wallet_exit, our_exit = trade['price'], get_current_price(trade['asset']) or trade['price']
                wallet_pnl = (wallet_exit - pos['wallet_entry']) * pos['wallet_shares']
                our_pnl    = (our_exit    - pos['our_entry'])    * pos['our_shares']
                paper_balance += pos['our_usd_in'] + our_pnl

                closed_trades.append({**pos, 'closed_at': now_str, 'wallet_exit': wallet_exit, 'our_exit': our_exit, 'wallet_pnl': wallet_pnl, 'our_pnl': our_pnl, 'wallet_roi': (wallet_exit-pos['wallet_entry'])/pos['wallet_entry'], 'our_roi': (our_exit-pos['our_entry'])/pos['our_entry']})
                events += 1
                alert_exit(pos, wallet_exit, our_exit)
                update_excel(open_positions, closed_trades, paper_balance, watch_list)
        time.sleep(0.3)

    total_pnl = sum(t['our_pnl'] for t in closed_trades)
    print(f"\r[#{cycle:04d} | {now_str}] Bal: ${paper_balance:,.2f} | Open: {len(open_positions)} | Closed: {len(closed_trades)} | P&L: ${total_pnl:+,.2f}", end='', flush=True)
    time.sleep(POLL_INTERVAL)