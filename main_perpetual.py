"""
Polymarket Copy Trade P&L Tracker
Railway deployment — runs 24/7 with State Persistence
"""
import requests, time, os, json, base64
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════
# CONFIG — edit these or set as Railway environment variables
# ═══════════════════════════════════════════════════════════
POLL_INTERVAL       = int(os.getenv('POLL_INTERVAL', 60))
LEADERBOARD_REFRESH = int(os.getenv('LEADERBOARD_REFRESH', 30))
TOP_N               = int(os.getenv('TOP_N', 100))
MIN_POSITION_USD    = float(os.getenv('MIN_POSITION_USD', 10))
PAPER_BALANCE       = float(os.getenv('PAPER_BALANCE', 1000.0))
BET_SIZE_PCT        = float(os.getenv('BET_SIZE_PCT', 0.05))
EXCEL_FILE          = os.getenv('EXCEL_FILE', 'copy_trade_pnl.xlsx')
STATE_FILE          = 'tracker_state.json'

DATA_API  = 'https://data-api.polymarket.com/v1'
CLOB_API  = 'https://clob.polymarket.com'
CATEGORIES = ['POLITICS', 'ECONOMICS', 'FINANCE', 'CULTURE', 'TECH', 'SPORTS', 'CRYPTO']

# GitHub config
GITHUB_TOKEN = os.getenv("GITHUB_TOKEN")
GITHUB_REPO  = os.getenv("GITHUB_REPO")

session = requests.Session()
session.headers.update({
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'application/json',
})

def log(msg):
    print(f'[{datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")}] {msg}', flush=True)

# ═══════════════════════════════════════════════════════════
# STATE RECOVERY & GITHUB UPLOAD
# ═══════════════════════════════════════════════════════════

def save_state(paper_balance, open_positions, closed_trades):
    state = {
        'paper_balance': paper_balance,
        'open_positions': open_positions,
        'closed_trades': closed_trades
    }
    try:
        with open(STATE_FILE, 'w') as f:
            json.dump(state, f)
    except Exception as e:
        log(f'⚠️ Failed to save state locally: {e}')

def load_state():
    # Try to pull from GitHub first to survive Railway restarts
    if GITHUB_TOKEN and GITHUB_REPO:
        log("🌐 Checking GitHub for previous memory state...")
        url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{STATE_FILE}"
        headers = {
            "Authorization": f"token {GITHUB_TOKEN}",
            "Accept": "application/vnd.github.v3+json"
        }
        try:
            res = requests.get(url, headers=headers)
            if res.status_code == 200:
                content = base64.b64decode(res.json().get("content", "")).decode('utf-8')
                with open(STATE_FILE, 'w') as f:
                    f.write(content)
                log("☁️ Successfully downloaded memory state from GitHub!")
        except Exception as e:
            pass # Normal if it's the very first time running

    # Load local file
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, 'r') as f:
                state = json.load(f)
            log(f"💾 Memory recovered! Restored {len(state.get('open_positions',{}))} open, {len(state.get('closed_trades',[]))} closed trades.")
            return state
        except Exception as e:
            log(f"⚠️ Could not read state file: {e}")
    return None

def upload_to_github(file_path, commit_msg="Hourly Sync"):
    if not GITHUB_TOKEN or not GITHUB_REPO:
        return

    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{file_path}"
    try:
        with open(file_path, "rb") as f:
            content = base64.b64encode(f.read()).decode()

        headers = {
            "Authorization": f"token {GITHUB_TOKEN}",
            "Accept": "application/vnd.github.v3+json"
        }

        res = requests.get(url, headers=headers)
        sha = res.json().get("sha") if res.status_code == 200 else None

        data = {
            "message": commit_msg,
            "content": content,
            "branch": "main"
        }
        if sha:
            data["sha"] = sha

        res = requests.put(url, headers=headers, json=data)

        if res.status_code in [200, 201]:
            log(f"☁️ {file_path} successfully synced to GitHub")
        else:
            log(f"❌ GitHub upload failed for {file_path}")
    except Exception as e:
        log(f"❌ Upload error for {file_path}: {e}")

# ═══════════════════════════════════════════════════════════
# LEADERBOARD
# ═══════════════════════════════════════════════════════════

def fetch_top_wallets(top_n=TOP_N):
    wallet_map = {}
    for cat in CATEGORIES:
        try:
            res = session.get(
                f'{DATA_API}/leaderboard',
                params={'timePeriod': 'MONTH', 'limit': top_n, 'orderBy': 'PNL', 'category': cat},
                timeout=15
            )
            if res.status_code != 200:
                continue
            traders = res.json()
            if not isinstance(traders, list):
                continue
            for t in traders:
                w = t.get('proxyWallet', '')
                if not w or not w.startswith('0x'):
                    continue
                pnl = float(t.get('pnl', 0))
                if w not in wallet_map:
                    wallet_map[w] = {
                        'username':   t.get('userName') or w[:10],
                        'wallet':     w,
                        'categories': [cat],
                        'pnl':        pnl,
                    }
                else:
                    if cat not in wallet_map[w]['categories']:
                        wallet_map[w]['categories'].append(cat)
                    wallet_map[w]['pnl'] = max(wallet_map[w]['pnl'], pnl)
            time.sleep(0.3)
        except Exception as e:
            log(f'⚠️ Leaderboard error ({cat}): {e}')

    for w in wallet_map:
        wallet_map[w]['category'] = '+'.join(wallet_map[w]['categories'])
    return wallet_map

def build_watch_list(live_wallets, open_positions, seen_tx):
    combined = dict(live_wallets)
    locked_wallets = {pos['wallet'] for pos in open_positions.values()}
    for wallet in locked_wallets:
        if wallet not in combined:
            pos = next(p for p in open_positions.values() if p['wallet'] == wallet)
            combined[wallet] = {
                'username': pos['username'],
                'wallet':   wallet,
                'category': pos['category'],
                'pnl':      0,
                'locked':   True,
            }
    for wallet in combined:
        seen_tx.setdefault(wallet, set())
    return combined

# ═══════════════════════════════════════════════════════════
# API HELPERS
# ═══════════════════════════════════════════════════════════

def get_activity(wallet, limit=20):
    try:
        res = session.get(f'{DATA_API}/activity',
                         params={'user': wallet, 'limit': limit}, timeout=15)
        return res.json() if res.status_code == 200 else []
    except:
        return []

def get_current_price(token_id):
    try:
        res = session.get(f'{CLOB_API}/midpoint',
                         params={'token_id': token_id}, timeout=10)
        if res.status_code == 200:
            return float(res.json().get('mid', 0))
    except:
        pass
    return None

def parse_trade(item):
    try:
        price = float(item.get('price', 0))
        size  = float(item.get('size', 0))
        return {
            'market':       item.get('title') or item.get('market', 'Unknown'),
            'outcome':      item.get('outcome', '?'),
            'side':         item.get('side', '?'),
            'price':        round(price, 4),
            'prob_pct':     round(price * 100, 1),
            'size':         round(size, 2),
            'usd_value':    round(price * size, 2),
            'tx_hash':      item.get('transactionHash', ''),
            'timestamp':    item.get('timestamp', 0),
            'asset':        item.get('asset', ''),
            'condition_id': item.get('conditionId', ''),
            'outcome_index':item.get('outcomeIndex', -1),
        }
    except:
        return None

def pos_key(wallet, condition_id, outcome_index):
    return f'{wallet[:10]}_{condition_id[:12]}_{outcome_index}'

# ═══════════════════════════════════════════════════════════
# EXCEL
# ═══════════════════════════════════════════════════════════

OPEN_COLS = [
    'Opened At', 'Trader', 'Category', 'Market', 'Outcome',
    'Wallet Entry', 'Our Entry', 'Entry Slip',
    'Wallet Shares', 'Our Shares', 'Our USD In',
    'Status', 'Asset Token', 'Wallet'
]
CLOSED_COLS = [
    'Opened At', 'Closed At', 'Trader', 'Category', 'Market', 'Outcome',
    'Wallet Entry', 'Our Entry', 'Entry Slip',
    'Wallet Exit',  'Our Exit',  'Exit Slip',
    'Wallet P&L $', 'Our P&L $', 'P&L Diff $', 'P&L Diff %',
    'Wallet ROI %', 'Our ROI %',
    'Our Shares', 'Our USD In', 'Result'
]

def _hdr(ws, cols, bg='0F3460'):
    thin = Side(style='thin', color='CCCCCC')
    b = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = b
    ws.row_dimensions[1].height = 36

def _row(ws, r, vals, bg='FFFFFF', pnl_cols=None):
    thin = Side(style='thin', color='CCCCCC')
    b = Border(left=thin, right=thin, top=thin, bottom=thin)
    pnl_cols = pnl_cols or []
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        if i in pnl_cols and isinstance(v, (int, float)):
            c.font = Font(name='Arial', size=9, bold=True,
                         color='1a6b2e' if v >= 0 else '8b0000')
        else:
            c.font = Font(name='Arial', size=9)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = b
    ws.row_dimensions[r].height = 16

def init_excel():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Open Trades'
    _hdr(ws1, OPEN_COLS)
    for i, w in enumerate([20,16,20,55,10,12,12,10,12,10,11,14,42,42], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = 'A2'

    ws2 = wb.create_sheet('Closed Trades')
    _hdr(ws2, CLOSED_COLS)
    for i, w in enumerate([20,20,14,18,55,10,10,10,10,10,10,10,13,13,13,12,11,11,10,11,10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'

    ws3 = wb.create_sheet('Summary')
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 22

    ws4 = wb.create_sheet('Wallet Log')
    _hdr(ws4, ['Refreshed At', 'Wallet', 'Username', 'Category', 'PnL Month $', 'Status'], bg='16213E')
    for i, w in enumerate([22, 44, 20, 22, 16, 14], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.freeze_panes = 'A2'

    wb.save(EXCEL_FILE)

def update_excel(open_positions, closed_trades, paper_balance, watch_list):
    try:
        wb = load_workbook(EXCEL_FILE)

        ws1 = wb['Open Trades']
        for row in ws1.iter_rows(min_row=2):
            for cell in row: cell.value = None
        for i, (key, pos) in enumerate(open_positions.items(), 2):
            slip   = round(pos['our_entry'] - pos['wallet_entry'], 4)
            locked = pos.get('locked', False)
            _row(ws1, i, [
                pos['opened_at'], pos['username'], pos['category'],
                pos['market'], pos['outcome'],
                pos['wallet_entry'], pos['our_entry'], slip,
                pos['wallet_shares'], pos['our_shares'], pos['our_usd_in'],
                'LOCKED (off LB)' if locked else 'Active',
                pos.get('asset',''), pos['wallet']
            ], bg='FFF9C4' if not locked else 'FFE0B2')

        ws2 = wb['Closed Trades']
        for row in ws2.iter_rows(min_row=2):
            for cell in row: cell.value = None
        for i, t in enumerate(closed_trades, 2):
            bg = 'E8F5E9' if t['our_pnl'] >= 0 else 'FFEBEE'
            _row(ws2, i, [
                t['opened_at'], t['closed_at'],
                t['username'], t['category'], t['market'], t['outcome'],
                t['wallet_entry'], t['our_entry'],
                round(t['our_entry'] - t['wallet_entry'], 4),
                t['wallet_exit'], t['our_exit'],
                round(t['our_exit'] - t['wallet_exit'], 4),
                round(t['wallet_pnl'], 2), round(t['our_pnl'], 2),
                round(t['our_pnl'] - t['wallet_pnl'], 2),
                round((t['our_pnl']-t['wallet_pnl'])/abs(t['wallet_pnl'])*100,1) if t['wallet_pnl'] else 0,
                round(t['wallet_roi']*100, 2), round(t['our_roi']*100, 2),
                t['our_shares'], t['our_usd_in'],
                'WIN' if t['our_pnl'] > 0 else 'LOSS'
            ], bg=bg, pnl_cols=[13, 14, 15])

        ws3 = wb['Summary']
        for row in ws3.iter_rows():
            for cell in row: cell.value = None
        n            = len(closed_trades)
        total_our    = sum(t['our_pnl']    for t in closed_trades)
        total_wallet = sum(t['wallet_pnl'] for t in closed_trades)
        wins         = sum(1 for t in closed_trades if t['our_pnl'] > 0)
        avg_se = sum(t['our_entry']-t['wallet_entry'] for t in closed_trades)/n if n else 0
        avg_sx = sum(t['our_exit'] -t['wallet_exit']  for t in closed_trades)/n if n else 0
        locked_count = sum(1 for p in open_positions.values() if p.get('locked'))
        rows = [
            ('SUMMARY', ''),
            ('Updated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
            ('', ''),
            ('Paper Balance',    f'${paper_balance:,.2f}'),
            ('Wallets Watching', len(watch_list)),
            ('Locked (off LB)',  locked_count),
            ('Open Positions',   len(open_positions)),
            ('Closed Trades',    n),
            ('', ''),
            ('Our Total P&L',    f'${total_our:+,.2f}'),
            ('Wallet Total P&L', f'${total_wallet:+,.2f}'),
            ('Slippage Cost',    f'${total_our-total_wallet:+,.2f}'),
            ('', ''),
            ('Win Rate',         f'{wins/n*100:.1f}%' if n else 'N/A'),
            ('Wins / Losses',    f'{wins} / {n-wins}'),
            ('', ''),
            ('Avg Entry Slip',   f'{avg_se*100:+.2f}c'),
            ('Avg Exit Slip',    f'{avg_sx*100:+.2f}c'),
        ]
        for r, (lbl, val) in enumerate(rows, 1):
            ws3.cell(r, 1, lbl).font = Font(name='Arial', bold=(r==1), size=10)
            ws3.cell(r, 2, val).font = Font(name='Arial', size=10)

        wb.save(EXCEL_FILE)
    except Exception as e:
        log(f'⚠️ Excel error: {e}')

def log_wallet_refresh(watch_list):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws4 = wb['Wallet Log']
        now = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
        r = ws4.max_row + 1
        for info in watch_list.values():
            _row(ws4, r, [
                now, info['wallet'], info['username'],
                info['category'], round(info['pnl'], 2),
                'LOCKED' if info.get('locked') else 'LIVE'
            ], bg='FFF3CD' if info.get('locked') else 'F0F8FF')
            r += 1
        wb.save(EXCEL_FILE)
    except Exception as e:
        log(f'⚠️ Wallet log error: {e}')

# ═══════════════════════════════════════════════════════════
# MAIN LOOP
# ═══════════════════════════════════════════════════════════

def main():
    global paper_balance

    log('🚀 Polymarket Copy Trade Tracker starting...')
    
    # 1. Ensure Excel file exists structurally
    init_excel()

    seen_tx        = {}
    watch_list     = {}

    # 2. Attempt to recover memory state
    state = load_state()
    if state:
        paper_balance  = state.get('paper_balance', PAPER_BALANCE)
        open_positions = state.get('open_positions', {})
        closed_trades  = state.get('closed_trades', [])
    else:
        paper_balance  = PAPER_BALANCE
        open_positions = {}
        closed_trades  = []

    log(f'   Balance: ${paper_balance:,.0f} | Bet: {BET_SIZE_PCT*100:.0f}% | Top {TOP_N}/cat')

    log('🌐 Fetching initial leaderboard...')
    watch_list = fetch_top_wallets(TOP_N)
    watch_list = build_watch_list(watch_list, open_positions, seen_tx)
    log(f'   Got {len(watch_list)} unique wallets')

    # 3. Immediately rebuild Excel with recovered data
    update_excel(open_positions, closed_trades, paper_balance, watch_list)

    log('📸 Snapshotting existing activity (Ignoring past trades to avoid double-counting)...')
    for wallet, info in watch_list.items():
        activity = get_activity(wallet, limit=50)
        seen_tx[wallet] = {item.get('transactionHash','') for item in activity if item.get('transactionHash')}
        time.sleep(0.3)

    log_wallet_refresh(watch_list)
    log(f'✅ Ready. Watching {len(watch_list)} wallets. Starting monitor loop...\n')

    cycle = 0
    last_github_push = time.time()

    while True:
        cycle += 1
        now_str = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
        events  = 0

        # Refresh leaderboard
        if cycle % LEADERBOARD_REFRESH == 0:
            log(f'🔄 Refreshing leaderboard (cycle {cycle})...')
            prev = set(watch_list.keys())
            fresh = fetch_top_wallets(TOP_N)
            watch_list = build_watch_list(fresh, open_positions, seen_tx)
            log_wallet_refresh(watch_list)

        # Poll wallets
        for wallet, info in watch_list.items():
            activity = get_activity(wallet, limit=20)

            for item in activity:
                tx   = item.get('transactionHash', '')
                side = item.get('side', '')
                if not tx or tx in seen_tx.get(wallet, set()):
                    continue
                seen_tx.setdefault(wallet, set()).add(tx)

                trade = parse_trade(item)
                if not trade or trade['usd_value'] < MIN_POSITION_USD:
                    continue

                key = pos_key(wallet, trade['condition_id'], trade['outcome_index'])

                # ENTRY
                if side == 'BUY' and key not in open_positions:
                    our_price  = get_current_price(trade['asset']) or trade['price']
                    bet_usd    = round(paper_balance * BET_SIZE_PCT, 2)
                    our_shares = round(bet_usd / our_price, 2) if our_price > 0 else 0
                    paper_balance -= bet_usd
                    slip = round((our_price - trade['price']) * 100, 2)

                    open_positions[key] = {
                        'username':      info['username'],
                        'wallet':        wallet,
                        'category':      info['category'],
                        'market':        trade['market'],
                        'outcome':       trade['outcome'],
                        'opened_at':     now_str,
                        'wallet_entry':  trade['price'],
                        'our_entry':     our_price,
                        'wallet_shares': trade['size'],
                        'our_shares':    our_shares,
                        'our_usd_in':    bet_usd,
                        'asset':         trade['asset'],
                        'condition_id':  trade['condition_id'],
                        'outcome_index': trade['outcome_index'],
                        'locked':        info.get('locked', False),
                    }
                    events += 1
                    log(f'🟢 ENTRY | {info["username"]:<18} | {trade["outcome"]:<5} | '
                        f'Wallet:{trade["price"]:.3f} Ours:{our_price:.3f} Slip:{slip:+.2f}c | '
                        f'${bet_usd:.2f} | {trade["market"][:50]}')
                    
                    update_excel(open_positions, closed_trades, paper_balance, watch_list)
                    save_state(paper_balance, open_positions, closed_trades)

                # EXIT
                elif side == 'SELL' and key in open_positions:
                    pos         = open_positions.pop(key)
                    wallet_exit = trade['price']
                    our_exit    = get_current_price(trade['asset']) or trade['price']
                    wallet_pnl  = (wallet_exit - pos['wallet_entry']) * pos['wallet_shares']
                    our_pnl     = (our_exit    - pos['our_entry'])    * pos['our_shares']
                    paper_balance += pos['our_usd_in'] + our_pnl

                    closed_trades.append({
                        **pos,
                        'closed_at':   now_str,
                        'wallet_exit': wallet_exit,
                        'our_exit':    our_exit,
                        'wallet_pnl':  wallet_pnl,
                        'our_pnl':     our_pnl,
                        'wallet_roi':  (wallet_exit-pos['wallet_entry'])/pos['wallet_entry'],
                        'our_roi':     (our_exit-pos['our_entry'])/pos['our_entry'],
                    })
                    events += 1
                    result = 'WIN' if our_pnl >= 0 else 'LOSS'
                    log(f'{"✅" if our_pnl>=0 else "❌"} EXIT  | {pos["username"]:<18} | {pos["outcome"]:<5} | '
                        f'Wallet P&L:{wallet_pnl:+.2f} Ours:{our_pnl:+.2f} | '
                        f'{result} | Bal:${paper_balance:,.2f}')
                    
                    update_excel(open_positions, closed_trades, paper_balance, watch_list)
                    save_state(paper_balance, open_positions, closed_trades)

            time.sleep(0.3)

        # Status
        total_pnl = sum(t['our_pnl'] for t in closed_trades)
        wins      = sum(1 for t in closed_trades if t['our_pnl'] > 0)
        locked_n  = sum(1 for w in watch_list.values() if w.get('locked'))
        log(f'[#{cycle:04d}] Bal:${paper_balance:,.2f} | '
            f'Watch:{len(watch_list)}({locked_n}locked) | '
            f'Open:{len(open_positions)} | '
            f'Closed:{len(closed_trades)} W{wins}/L{len(closed_trades)-wins} | '
            f'P&L:${total_pnl:+,.2f}'
            + (f' | ⚡{events}' if events else ''))

        # 🔥 Time-based hourly push to GitHub (3600 seconds)
        if time.time() - last_github_push >= 3600:
            log("⏳ 1 hour elapsed, triggering GitHub sync...")
            upload_to_github(EXCEL_FILE, "Hourly Excel Update")
            upload_to_github(STATE_FILE, "Hourly Memory Backup")
            last_github_push = time.time()

        time.sleep(POLL_INTERVAL)

if __name__ == '__main__':
    main()
